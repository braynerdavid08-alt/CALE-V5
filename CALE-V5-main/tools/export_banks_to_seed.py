# -*- coding: utf-8 -*-
"""Export CALE Excel banks to JSON seed files with quality fixes."""
import json
import re
from pathlib import Path
import openpyxl

OUT = Path(r"d:\APPS programacion\CALE-V5-main\CALE-V5-main\src\Cale.Api\SeedData")
OUT.mkdir(parents=True, exist_ok=True)


def clean_normas_stem(text: str) -> str:
    """Remove the embedded answer leak ('Regla aplicable: ...') from situational stems."""
    text = (text or "").strip()
    m = re.search(r"\n*\s*Regla aplicable\s*:", text, re.IGNORECASE)
    if m:
        text = text[: m.start()].strip()
    text = re.sub(r"\n{3,}", "\n\n", text).strip()
    return text


def export_normas():
    path = r"d:\APPS programacion\CALE_Banco_500_Normas_Transito.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Banco Normas"]
    rows = list(ws.iter_rows(values_only=True))
    questions = []
    stripped = 0
    leaks = 0
    for r in rows[1:]:
        if r[0] is None:
            continue
        raw = str(r[5] or "").strip()
        stem = clean_normas_stem(raw)
        if stem != raw:
            stripped += 1
        corr = str(r[10] or "").strip().upper()
        opts = [
            {"text": str(r[6] or "").strip(), "isCorrect": corr == "A"},
            {"text": str(r[7] or "").strip(), "isCorrect": corr == "B"},
            {"text": str(r[8] or "").strip(), "isCorrect": corr == "C"},
            {"text": str(r[9] or "").strip(), "isCorrect": corr == "D"},
        ]
        if sum(1 for o in opts if o["isCorrect"]) != 1:
            raise SystemExit(f"Bad correct letter at ID {r[0]}: {corr}")
        if any(not o["text"] for o in opts) or not stem:
            raise SystemExit(f"Empty fields at ID {r[0]}")
        correct_text = next(o["text"] for o in opts if o["isCorrect"])
        if correct_text in stem:
            leaks += 1
        tipo = str(r[3] or "").strip()
        qtype = (
            "Verdadero/Falso"
            if "Verdadero" in tipo or "Falso" in tipo
            else "Seleccion multiple"
        )
        questions.append(
            {
                "externalId": int(r[0]),
                "subject": str(r[1] or "").strip(),
                "topic": str(r[2] or "").strip(),
                "subtopic": str(r[2] or "").strip(),
                "difficulty": str(r[4] or "").strip(),
                "type": qtype,
                "text": stem,
                "explanation": str(r[11] or "").strip(),
                "source": str(r[12] or "").strip(),
                "options": [
                    {"text": o["text"], "isCorrect": o["isCorrect"]} for o in opts
                ],
            }
        )

    payload = {
        "bankName": "Normas de tránsito (Colombia)",
        "description": "Banco de 500 preguntas situacionales y de verdadero/falso sobre normas de tránsito colombianas (Ley 769/2002, RUCT y Manual de Señalización Vial 2024). Versión CALE-NT-2026-01.",
        "blockName": "Normas de tránsito",
        "questions": questions,
    }
    out = OUT / "banco-normas-transito.json"
    out.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    print(
        f"Normas: {len(questions)} questions, stems cleaned: {stripped}, "
        f"remaining answer leaks: {leaks} -> {out}"
    )


def export_senales():
    path = r"d:\APPS programacion\CALE_Banco_Senales_Reconocimiento_Visual.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Reconocimiento Señales"]
    rows = list(ws.iter_rows(values_only=True))
    questions = []
    for r in rows[1:]:
        if r[0] is None:
            continue
        code = str(r[2] or "").strip()
        corr = str(r[8] or "").strip().upper()
        opts = [
            {"text": str(r[4] or "").strip(), "isCorrect": corr == "A"},
            {"text": str(r[5] or "").strip(), "isCorrect": corr == "B"},
            {"text": str(r[6] or "").strip(), "isCorrect": corr == "C"},
            {"text": str(r[7] or "").strip(), "isCorrect": corr == "D"},
        ]
        if sum(1 for o in opts if o["isCorrect"]) != 1:
            raise SystemExit(f"Bad signal correct at {r[0]}: {corr}")
        questions.append(
            {
                "externalId": str(r[0]).strip(),
                "subject": str(r[1] or "").strip(),
                "topic": code,
                "subtopic": str(r[1] or "").strip(),
                "difficulty": str(r[9] or "").strip(),
                "type": "Seleccion multiple",
                "text": str(r[3] or "").strip(),
                "imageUrl": f"/signals/{code}.svg",
                "explanation": (
                    f"Señal {code}: la denominación oficial según el Manual de "
                    "Señalización Vial de Colombia 2024 es la opción marcada como correcta."
                ),
                "source": str(r[11] or "").strip(),
                "options": opts,
            }
        )

    payload = {
        "bankName": "Reconocimiento visual de señales",
        "description": "Banco de 194 preguntas de reconocimiento de señales verticales (reglamentarias, preventivas e informativas) según el Manual de Señalización Vial de Colombia 2024.",
        "blockName": "Señalización vial",
        "questions": questions,
    }
    out = OUT / "banco-senales-reconocimiento.json"
    out.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    print(f"Señales: {len(questions)} questions -> {out}")

    cat = []
    for q in questions:
        correct = next(o["text"] for o in q["options"] if o["isCorrect"])
        cat.append(
            {
                "code": q["topic"],
                "family": q["subject"],
                "name": correct,
                "imageUrl": q["imageUrl"],
            }
        )
    (OUT / "senales-catalog.json").write_text(
        json.dumps(cat, ensure_ascii=False, indent=2), encoding="utf-8"
    )


if __name__ == "__main__":
    export_normas()
    export_senales()
